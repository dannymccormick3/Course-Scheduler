import re
from collections import namedtuple
from openpyxl import load_workbook
import time
import pprint

def create_course_dict():
    """
    Creates a dictionary containing course info.
    Keys: namedtuple of the form ('program, designation')
    Values: namedtuple of the form('name, prereqs, credits')
            prereqs is a tuple of prereqs where each prereq has the same form as the keys
    """
    wb = load_workbook('newcatalog.xlsx')
    catalog = wb.get_sheet_by_name('catalog')
    Course = namedtuple('Course', 'program, designation')
    CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
    course_dict = {}
    for row in range(1, catalog.max_row + 1):
        key = Course(get_val(catalog, 'A', row), get_val(catalog, 'B', row))
        prereqs = tuple(tuple(get_split_course(prereq) for prereq in prereqs.split())
                   for prereqs in none_split(get_val(catalog, 'E', row)))
        val = CourseInfo(get_val(catalog, 'C', row), tuple(get_val(catalog, 'D', row).split()), prereqs)
        course_dict[key] = val
    return course_dict


def get_split_course(course):
    """
    Parses a course from programdesignation into the ('program, designation') form.
    e.g. 'CS1101' -> ('CS', '1101')
    """
    return tuple(split_course for course_part in re.findall('((?:[A-Z]+-)?[A-Z]+)(.+)', course)
                 for split_course in course_part)


def none_split(val):
    """Handles calling split on a None value by returning the empty list."""
    return val.split(', ') if val else ()


def get_val(catalog, col, row):
    """Returns the value of a cell."""
    return catalog[col + str(row)].value


def print_dict(dict):
    """Simply prints a dictionary's key and values line by line."""
    for key in dict:
        print(key, dict[key])

"""
All above code is from the Artificial Intelligence TA.
"""

def course_scheduler (course_descriptions, goal_conditions, initial_state):
	"""
	PRE:
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		goal_conditions: List of post-conditions after the courses have been taken. 
			e.g. [(‘CS’, ‘mathematics’), (‘CS’, ‘core’), (‘MATH’, ’3641’), (‘CS’, ’1151’), (‘MATH’, ‘2410’)]
		initial_state: List of pre-condition courses that have already been credited to the student.
			e.g. [('CS', '1101'), ('SPAN', '1101')]
	POST:
		Returns a set of scheduled courses that satisfy all of the goal conditions.
			Scheduled courses are of the form (Course, scheduled term, number of credits)
			e.g. ((“CS”, “2201”), (“Spring”, “Frosh”), 3)
		Each semester must have 12-18 hours scheduled. If no such satisfying set exists (in 4 years of semesters using
		only fall and spring terms), returns an empty set.
	"""
	goal_conditions_map = add_semester_field(goal_conditions, 8)
	initial_state_map = add_semester_field(initial_state, 0)
	tot_hours_per_semester = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0}
	schedule, tot_hours_per_semester = create_satisfying_schedule(course_descriptions, goal_conditions_map, initial_state_map, tot_hours_per_semester)
	schedule = fill_courseload(course_descriptions, schedule, tot_hours_per_semester)
	return format_schedule(schedule, course_descriptions)

def format_schedule(schedule, course_descriptions):
	# TODO
	formatted_schedule = {}
	for i in range (1,9):
		for course in schedule:
			if schedule[course] == i:
				semester_number = schedule[course]
				semester = "Spring"
				year = "Frosh"
				if semester_number % 2 == 1:
					semester = "Fall"
					semester_number += 1
				semester_number = semester_number / 2
				if semester_number == 2:
					year = "Sophomore"
				elif semester_number == 3:
					year = "Junior"
				elif semester_number == 4:
					year = "Senior"
				formatted_schedule[(course[0], course[1])] = (course_descriptions[course][0], (semester, year), ())
	return formatted_schedule

def create_satisfying_schedule (course_descriptions, goal_conditions, completed_courses, tot_hours_per_semester):
	"""
	PRE:
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		goal_conditions: Map mapping courses to be taken to the latest semester in which they can be taken given the
			current set of completed courses. Numbers are used to represent the semesters (1="Fall freshman", 2="Spring freshman"...)
			e.g. {(‘CS’, ‘mathematics’): 8, (‘CS’, ‘core’): 8, (‘MATH’, ’3641’): 7, (‘CS’, ’1151’): 7, (‘MATH’, ‘2410’): 3}
		completed_courses: Map mapping scheduled courses to the semester they are to be completed. A zero indicates they are pre-conditions
			e.g. {('CS', '1101'): 1, ('SPAN', '1101'): 0, ('CS', '2201'): 2}
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
	POST:
		Returns 2 dictionaries. The first maps courses to the semester that they will be taken. The second maps semesters to the total number
			of hours taken that semester. All prerequisites are satisfied and no semester has more than 18 hours.
	"""
	if len(goal_conditions) == 0:
		return completed_courses, tot_hours_per_semester
	for goal_class in goal_conditions:
		goal_latest_semester = goal_conditions[goal_class]
		if goal_class in completed_courses:
			# If the current goal class has already been assigned and satisfies the semester requirement, skip this goal. Otherwise remove it from the assigned classes.
			# TODO: Make this more efficient (so it doesn't redo search no previously searched paths) - maybe do it in process_course? - I actually think that's why this is failing
			if completed_courses[goal_class] <= goal_latest_semester:
				new_goal_conditions = dict(goal_conditions)
				del new_goal_conditions[goal_class]
				return create_satisfying_schedule(course_descriptions, new_goal_conditions, completed_courses, tot_hours_per_semester)
			else:
				tot_hours_per_semester[completed_courses[goal_class]] -= int(course_descriptions[goal_class][0])
				del completed_courses[goal_class]
		perform_op = process_course (course_descriptions, dict(goal_conditions), dict(completed_courses), dict(tot_hours_per_semester), goal_class)
		if len(perform_op) > 0:
			return perform_op
	return {}

def process_course (course_descriptions, goal_conditions, completed_courses, tot_hours_per_semester, course):
	"""
	PRE:
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		goal_conditions: Map mapping courses to be taken to the latest semester in which they can be taken given the
			current set of completed courses. Numbers are used to represent the semesters (1="Fall freshman", 2="Spring freshman"...)
			e.g. {(‘CS’, ‘mathematics’): 8, (‘CS’, ‘core’): 8, (‘MATH’, ’3641’): 7, (‘CS’, ’1151’): 7, (‘MATH’, ‘2410’): 3}
		completed_courses: Map mapping scheduled courses to the semester they are to be completed. A zero indicates they are pre-conditions
			e.g. {('CS', '1101'): 1, ('SPAN', '1101'): 0, ('CS', '2201'): 2}
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
		course: Tuple of the course to be processed.
	POST:
		Determines if the given course can be added to the current set of completed_courses without violating hours requirements/its latest 
			semester as specified in goal conditions. If it can't returns an empty dictionary. If it can, adds it to the schedule, adds its 
			prereqs to goal conditions and returns the result of create_satisfying_schedule with the updated dictionaries. If there are multiple
			sets of allowable prereqs, tries all sets until one returns a non-empty dictionary or all possibilities are exhausted.
	"""
	course_info = course_descriptions[course]
	credit_hours = int(course_info[0])
	semester = get_semester_assignment (int(goal_conditions[course]), tot_hours_per_semester, credit_hours, course_info[1])
	if semester == -1:
		# If no satisfying semester assignment, course can't be processed.
		return {}

	# Remove course from goal conditions, add to completed_courses, update hours_per_semester
	del goal_conditions[course]
	completed_courses[course] = semester
	tot_hours_per_semester[semester] += credit_hours

	# Add new goal conditions
	prereqs = course_info[2]
	pre_semester = semester
	if credit_hours > 0:
		# If non-zero number of credit hours, goals need to go in previous semester at the latest
		pre_semester -= 1
	if len(prereqs) == 0:
		# If no prereqs, just move to next goal
		return create_satisfying_schedule(course_descriptions, dict(goal_conditions), dict(completed_courses), dict(tot_hours_per_semester))
	for course_set in prereqs:
		# Seperately branch for each set of prereqs
		new_goal_conditions = dict(goal_conditions)
		for course in course_set:
			# If its already in goal_conditions, make last possible semester the smaller of the 2 values
			if course in new_goal_conditions:
				new_goal_conditions[course] = min(new_goal_conditions[course], pre_semester)
			else:
				new_goal_conditions[course] = pre_semester
		if len(new_goal_conditions) == 0:
			return completed_courses
		else:
			attempt = create_satisfying_schedule(course_descriptions, new_goal_conditions, dict(completed_courses), dict(tot_hours_per_semester))
			if len(attempt) != 0:
				return attempt
	return {}

def get_semester_assignment (latest_semester, tot_hours_per_semester, credit_hours, terms):
	"""
	PRE:
		latest_semester: Integer value of the latest semester a course may be scheduled
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
		credit_hours: Integer value of the number of credit hours that a course counts for.
		terms: Tuple that may contain 'Fall', 'Spring', and/or 'Summer' representing which terms a course can be taken in.
	POST:
		Returns the latest semester that that the course can be taken in without violating any of the constraints and ensuring
			that the number of hours taken in that semester remains less than or equal to 18
	"""
	i = latest_semester
	while i > 0:
		if tot_hours_per_semester[i] + credit_hours <= 18:
			if i % 2 == 0:
				if 'Spring' in terms:
					return i
			elif 'Fall' in terms:
				return i
		i -= 1
	return -1

def add_semester_field (goal_conditions, semester):
	"""
	PRE:
		goal_conditions: Set of post conditions after all courses have been taken
		semester: The latest semester the goal can be filled
	POST:
		Returns map of post conditions after all courses have been taken mapped to the latest semester they can be fulfilled
		e.g. {(‘CS’, ‘mathematics’): 8 (‘CS’, ‘core’): 8 (‘MATH’, ’3641’): 8 (‘CS’, ’1151’): 8 (‘MATH’, ‘2410’): 8}
	""" 
	updated_conditions = {}
	for condition in goal_conditions:
		updated_conditions[condition] = semester
	return updated_conditions

def fill_courseload (course_descriptions, schedule, tot_hours_per_semester):
	"""
	PRE:
		course_descriptions: Dictionary of courses with the following form.
			key - Course(program='<department>', designation='<number>') e.g. Course(program='CS', designation='2201')
			value - CourseInfo(credits='<# credit hours>', terms=<terms list>, prereqs=<prereqs list>) 
				e.g. CourseInfo(credits='3', terms=('Spring', 'Fall'), prereqs=((('CS', '1101'),),))
		schedule: Map mapping scheduled courses to the semester they are to be completed. The schedule may be incomplete
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
			e.g. {1: 0, 2: 6, 3: 3, 4: 10, 5: 7, 6: 6, 7: 6, 8: 9}
	POST:
		Returns a valid schedule building off of the input schedule where all semesters have 12-18 hours.
	"""
	i = 1
	while tot_hours_per_semester[i] == 0 or tot_hours_per_semester[i] >= 12:
		i += 1
	while i < 9:
		for course in course_descriptions:
			course_info = course_descriptions[course]
			if is_valid_class(course, int(course_info[0]), course_info[1], course_info[2], schedule, tot_hours_per_semester, i):
				schedule[course] = i
				tot_hours_per_semester[i] += int(course_info[0])
				while tot_hours_per_semester[i] >= 12:
					i += 1
					if i == 9:
						return schedule
	return schedule

def is_valid_class (course, credit_hours, terms, prereqs, schedule, tot_hours_per_semester, semester):
	"""
	PRE:
		course: The course to determine the validity of.
		credit_hours: Integer representing the number of hours the course counts for.
		terms: Tuple that may contain 'Fall', 'Spring', and/or 'Summer' representing which terms a course can be taken in.
		prereqs: Tuple of tuples of satisfying prereqs to the course.
		schedule: Dictionary representing the current schedule.
		tot_hours_per_semester: Map mapping each semester (1-8) to the total number of hours taken in that semseter.
		semester: Integer representing the semester to assign the course.
	POST:
		Returns true if class can be assigned to that semester, false if it can't be.
	"""
	if course in schedule:
		return False
	if credit_hours + tot_hours_per_semester[semester] > 18:
		return False
	if semester % 2 == 0:
		if 'Spring' not in terms:
			return False
	elif 'Fall' not in terms:
		return False
	if len(prereqs) == 0:
		return True
	for req_set in prereqs:
		satisfying = True
		for req in prereqs:
			if req not in schedule or schedule[req] >= semester:
				satisfying = False
		if satisfying:
			return True
	return False

def main():
	# TODO: Add heuristic portion
	Course = namedtuple('Course', 'program, designation')
	goal_conditions = [Course('CS', 'mathematics'), Course('CS', 'core'), Course('MATH', '3641'), Course('CS', '1151'), Course('MATH', '2410')]
	initial_state = [Course('CS', '1101'), Course('SPAN', '1101')]
	plan = course_scheduler(create_course_dict(), goal_conditions, initial_state)
	pp = pprint.PrettyPrinter()

	pp.pprint (plan)

if __name__ == "__main__":
    main()